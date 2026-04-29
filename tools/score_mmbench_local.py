#!/usr/bin/env python3
"""
Local MMBench accuracy computation from the generated xlsx file.

The dev-split xlsx has both 'prediction' and 'answer' columns,
so no server submission is needed — accuracy is computed directly.

Usage:
    python TorchUMM/tools/score_mmbench_local.py \
        --xlsx TorchUMM/output/mmbench/janus_pro_quant_mbq_7b/mmbench_dev_20230712_260423024059.xlsx \
        [--out TorchUMM/output/mmbench/janus_pro_quant_mbq_7b/score.json]

    # Score all models at once:
    python TorchUMM/tools/score_mmbench_local.py --all
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not found. Run: pip install openpyxl")


def score_xlsx(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if "prediction" not in headers or "answer" not in headers:
        raise ValueError(f"xlsx missing 'prediction' or 'answer' column: {headers}")

    pred_idx = headers.index("prediction")
    ans_idx = headers.index("answer")

    total = correct = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        pred = str(row[pred_idx]).strip().upper() if row[pred_idx] is not None else ""
        ans = str(row[ans_idx]).strip().upper() if row[ans_idx] is not None else ""
        if not ans:
            continue
        total += 1
        if pred == ans:
            correct += 1

    acc = round(correct / total * 100, 2) if total else 0.0
    return {"correct": correct, "total": total, "accuracy_pct": acc}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=str, default=None)
    parser.add_argument("--out", type=str, default=None)
    parser.add_argument(
        "--all", action="store_true",
        help="Score all xlsx files found under TorchUMM/output/mmbench/",
    )
    args = parser.parse_args()

    repo = Path(__file__).resolve().parents[2]
    output_root = repo / "TorchUMM" / "output" / "mmbench"

    if args.all:
        xlsxs = sorted(output_root.rglob("mmbench_dev_*.xlsx"))
        if not xlsxs:
            sys.exit(f"No mmbench xlsx files found under {output_root}")
        results = {}
        for xp in xlsxs:
            model_dir = xp.parent.name
            res = score_xlsx(xp)
            results[model_dir] = res
            print(f"{model_dir}: {res['accuracy_pct']:.2f}%  ({res['correct']}/{res['total']})")
        # save summary
        out_path = output_root / "mmbench_local_scores.json"
        out_path.write_text(json.dumps(results, indent=2))
        print(f"\nSaved to {out_path}")
        return

    if not args.xlsx:
        parser.print_help()
        sys.exit(1)

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    res = score_xlsx(xlsx_path)
    print(f"Accuracy: {res['accuracy_pct']:.2f}%  ({res['correct']}/{res['total']})")

    if args.out:
        out_path = Path(args.out)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(res, indent=2))
        print(f"Saved to {out_path}")
    else:
        # auto-save next to the xlsx
        out_path = xlsx_path.parent / "mmbench_local_score.json"
        out_path.write_text(json.dumps(res, indent=2))
        print(f"Saved to {out_path}")


if __name__ == "__main__":
    main()
